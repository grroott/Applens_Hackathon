import openpyxl
obb=openpyxl.load_workbook("F://zxc.xlsx")
o=obb.active
col_name=[]
tp=o.max_row
o.cell(row=1,column=10).value='row_count'
j=1
for i in range(2,tp+1):
    o.cell(row=i,column=10).value=j
    j+=1
obb.save("F://zxc.xlsx")
for i in range(1,11):
    col_name.append(o.cell(row=1,column=i).value)

joi=" varchar(100), ".join(col_name)+" varchar(100)"

val_name=[]
val_names=[]
for j in range(2,11):
    val_name=[]
    for i in range(1,11):
        cx=str(o.cell(row=j, column=i).value)
        val_name.append(cx)
    val_name=tuple(val_name)
    val_names.append(val_name)

vari=(', '.join(['%s']*10))

import mysql.connector

mydb=mysql.connector.connect(host='localhost',user='root',passwd='root',database='test')

mycursor=mydb.cursor()
xx="create table if not exists zxc ({temp})"
mycursor.execute(xx.format(temp=joi))
mycursor.execute("ALTER TABLE zxc modify Work_Log varchar(6000)")
mycursor.execute("ALTER TABLE zxc modify Notes varchar(2000)")
mycursor.execute("ALTER TABLE zxc modify Support_Diary varchar(4000)")
mycursor.execute("ALTER TABLE zxc modify Resolution varchar(2000)")
up='''insert into zxc values (%s)''' %(vari)
mycursor.executemany(up, val_names)
mydb.commit()

#
#
# import mysql.connector
# from nltk.tokenize import sent_tokenize, word_tokenize
# import warnings
# import gensim
# from gensim.models import Word2Vec, KeyedVectors
# from nltk.corpus import stopwords
# from nltk.stem import PorterStemmer, WordNetLemmatizer
# import re
#
# warnings.filterwarnings(action='ignore')
# final_lst = []
# mydb=mysql.connector.connect(host='localhost',user='root',passwd='root',database='test')
#
#
# for iii in range(53, 54):
#     mycursor = mydb.cursor()
#     sql="select Assigned_to_Group,Item, Summary, Resolution, Support_Diary, Notes, Work_Log, 1CApp_Name from newtablee_e where row_count=(%s)"
#     form=(iii,)
#     mycursor.execute(sql,form)
#     pt = []
#     for i in mycursor:
#         for j in i:
#             try:
#                 j = j.replace(u'\xa0', u'')
#             except:
#                 pass
#             pt.append(j)
#
#     vip=[]
#     non_vip=[]
#     for b in pt:
#         if b == pt[0] or b == pt[7]:
#             vip.append(b)
#         else:
#             non_vip.append(b)
#     ptt = []
#     for i in non_vip:
#         x = i.split(" ")
#         ptt.append(x)
#
#     pttt = []
#     for i in ptt:
#         for j in i:
#             pttt.append(j)
#
#     lst = []
#     lstt = []
#
#     def Punctuation(string):
#         punctuations = '''!()-[]{};:'"\,<>'"./?@#$%^&*_~'''
#         for i in punctuations:
#             lst.append(i)
#         for i in pttt:
#             if i not in lst:
#                 lstt.append(i)
#
#     string = pttt
#     Punctuation(string)
#
#     for i in lstt:
#         if i == '':
#             lstt.remove(i)
#
#     strr = ''
#     for b in lstt:
#         strr = strr + " " + b
#
#     s = " ".join(strr.split())
#     example_sent = s
#     stop_words = set(stopwords.words('english'))
#
#     word_tokens = word_tokenize(example_sent)
#     filtered_sentence = []
#
#     for w in word_tokens:
#         if w not in stop_words:
#             filtered_sentence.append(w)
#     ls = []
#     lss = []
#     punctuations = '''!()-[]{};:'"\,<>'"./?@#$%^&*_~'''
#     for i in punctuations:
#         ls.append(i)
#     for i in filtered_sentence:
#         if i not in lst:
#             lss.append(i)
#     ps = PorterStemmer()
#
#     words = lss
#     words_aft_root = []
#     for w in words:
#         words_aft_root.append(ps.stem(w))
#
#     lemmatizer = WordNetLemmatizer()
#     words_aft_lem = []
#     for i in words_aft_root:
#         words_aft_lem.append(lemmatizer.lemmatize(i))
#     lower_case = []
#     for i in words_aft_lem:
#         lower_case.append(i.lower())
#     lower_cas = []
#     for i in lower_case:
#         pattern = re.compile(r'\d[:/]\d\d[:/]\d\d')
#         patternn=re.compile(r'\d\d\d\d\d\d')
#         matches = pattern.finditer(i)
#         matchess = patternn.finditer(i)
#         count = 0
#         for match in matches:
#             count += 1
#         for match in matchess:
#             count += 1
#         if i not in lower_cas:
#             if count == 0:
#                 lower_cas.append(i)
#     for i in vip:
#         lower_cas.append(i)
#
#     lower_ca=[]
#
#     for i in lower_cas:
#
#         pattern = re.compile(r'\d[:/]\d[:/]\d\d\d\d')
#         matches = pattern.finditer(i)
#         count = 0
#         for match in matches:
#             count += 1
#         if count == 0:
#             lower_ca.append(i)
#
#     final_lst.append(lower_ca)
# print(pt)
# print(vip)
# print(final_lst)
# model1=Word2Vec.load('thousand_sample_cbow_v2')
# model1.build_vocab(final_lst, update=True)
# model1.train(final_lst, epochs=model1.epochs, total_examples=model1.corpus_count)
# tmp=0
# for i in lower_ca:
#     t=model1.wv.get_vector(i)
# import numpy as np
# model_word_vector = np.array( t, dtype='f')
# prob=model1.most_similar([model_word_vector],[],topn=40000)
# app_na=['AVM PMO', 'AppLens', 'AppLensLite', 'AVM 1Capp', 'C2S', 'Genie Lamp', 'AVM DART', 'PriSize Estimation', 'Mini', 'Mobilo']
# for i in prob:
#     for j in i:
#         if j in app_na:
#             print(i)
#
# model2=Word2Vec.load('thousand_sample_skipgram_v2')
# bvb=model2.corpus_total_words
# model2.build_vocab(final_lst, update=True)
# model2.train(final_lst, epochs=model2.epochs, total_examples=model2.corpus_count)
# probb=model2.most_similar(positive=lower_ca,negative=[],topn=bvb)
# app_name=['AVM PMO', 'AppLens', 'AppLensLite', 'AVM 1Capp', 'C2S', 'Genie Lamp', 'AVM DART', 'PriSize Estimation', 'Mini', 'Mobilo']
#
#
# for i in probb:
#     for j in i:
#         if j in app_name:
#             print(i)
#
# print('c')
# model1=Word2Vec.load('thousand_sample_cbow_v2')
# bvb=model1.corpus_total_words
# model1.build_vocab(final_lst, update=True)
# model1.train(final_lst, epochs=model1.epochs, total_examples=model1.corpus_count)
# probb=model1.most_similar(positive=lower_ca,negative=[],topn=bvb)
# app_name=['AVM PMO', 'AppLens', 'AppLensLite', 'AVM 1Capp', 'C2S', 'Genie Lamp', 'AVM DART', 'PriSize Estimation', 'Mini', 'Mobilo']
#
#
# for i in probb:
#     for j in i:
#         if j in app_name:
#             print(i)
