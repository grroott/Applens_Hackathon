
import openpyxl
import pandas as pd
import numpy as np
import re
import pickle
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.externals import joblib

path = "E:\\1500to1600.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
for i in range(2, 3):
    x = ""
    final_lstt = []
    for j in range(1, 4):
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    final_lstt.append(x)
    print(final_lstt)
    # print(final_lstt)
    # final_lstt=[' AppLensLite_L2 AVM 1Capp Reports']
    # final_lstt=[' AVMPMO_L2 AVM PMO', ' AVMPMO_L2 AVM PMO',' AVMPMO_L2 AVM PMO', ' L1 (AVM Tools & Portals) AppLensLite']
    # final_lstt=[' AVMPMO_L2 AVM PMO', 'SDF',' AVMPMO_L2 AVM PMO', ' AVMPMO_L2 AVM PMO',' AVMPMO_L2 AVM PMO',' L1 (AVM Tools & Portals) AppLensLite']
    # print(len(final_lst))
    final_lstt = pd.DataFrame(final_lstt)
    final_lstt.columns = ["Details"]
    # nltk.download('stopwords')
    corpus = []
    for i in range(0, len(final_lstt)):
        text = re.sub('[^a-zA-Z]', '', final_lstt['Details'][i])
        text = text.lower()
        text = text.split()
        ps = PorterStemmer()
        text = ''.join(text)
        corpus.append(text)
    # creating bag of words model
    print(corpus)
    cvc = CountVectorizer(max_features=1500)
    XXX = cvc.fit_transform(corpus).toarray()
    print(XXX)
   # fitting naive bayes to the training set
    tmp = joblib.load('new_try.pickle')
    # f = open('saved.pickle', 'rb')
    # classifier = pickle.load(f)
    print(tmp.predict(XXX))



for i in range(3, 4):
    x = ""
    final_lstt = []
    for j in range(1, 4):
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    final_lstt.append(x)
    print(final_lstt)
    # print(final_lstt)
    # final_lstt=[' AppLensLite_L2 AVM 1Capp Reports']
    # final_lstt=[' AVMPMO_L2 AVM PMO', ' AVMPMO_L2 AVM PMO',' AVMPMO_L2 AVM PMO', ' L1 (AVM Tools & Portals) AppLensLite']
    # final_lstt=[' AVMPMO_L2 AVM PMO', 'SDF',' AVMPMO_L2 AVM PMO', ' AVMPMO_L2 AVM PMO',' AVMPMO_L2 AVM PMO',' L1 (AVM Tools & Portals) AppLensLite']
    # print(len(final_lst))
    final_lstt = pd.DataFrame(final_lstt)
    final_lstt.columns = ["Details"]
    # nltk.download('stopwords')
    corpus = []
    for i in range(0, len(final_lstt)):
        text = re.sub('[^a-zA-Z]', '', final_lstt['Details'][i])
        text = text.lower()
        text = text.split()
        ps = PorterStemmer()
        text = ''.join(text)
        corpus.append(text)
    # creating bag of words model
    print(corpus)
    cv = CountVectorizer(max_features=1500)
    XX = cv.fit_transform(corpus).toarray()
    print(XX)
   # fitting naive bayes to the training set
    tmp = joblib.load('new_try.pickle')
    # f = open('saved.pickle', 'rb')
    # classifier = pickle.load(f)
    print(tmp.predict(XX))






'''final_lstt=[' AVMPMO_L2 AVM PMO', ' AVMPMO_L2 AVM PMO',' AVMPMO_L2 AVM PMO', ' L1 (AVM Tools & Portals) AppLensLite']
for i in final_lstt:
    cc = pd.DataFrame(list(i))
    cc.columns = ["Details"]
# nltk.download('stopwords')
    corpus = []
    for j in range(0, len(cc)):
        text = re.sub('[^a-zA-Z]', '', cc['Details'][j])
        text = text.lower()
        text = text.split()
        ps = PorterStemmer()
        text = ''.join(text)
        corpus.append(text)
    print(corpus)
    st=""
    for k in corpus:
        st=st+k
    st=list(st)
# creating bag of words model
    cv = CountVectorizer(max_features=1500)
    X = cv.fit_transform(st).toarray()
# fitting naive bayes to the training set
    tmp = joblib.load('new_try.pickle')
# f = open('saved.pickle', 'rb')
# classifier = pickle.load(f)
    print(tmp.predict(X))'''

