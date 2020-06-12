# import necessary package/Library

import pandas as pd
import openpyxl
import re
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from sklearn.feature_extraction.text import CountVectorizer
import pickle
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import accuracy_score

# Read Excel for training the model

print("Reading Excel for training the model")
path = "D:\\Applens_Hackathon\\ticket_dump_for_training.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
final_lst = []
for i in range(2, 1000):
    x = ""
    lst = []
    for j in range(1, 4):
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    lst.append(x)
    for j in range(5, 6):
        cell_obj = sheet_obj.cell(row=i, column=j)
        lst.append(cell_obj.value)
    final_lst.append(lst)
final_lst = pd.DataFrame(final_lst)
final_lst.columns = ["Details", "Appname"]

# preparing efficient data for training

print("preparing efficient data for training")
corpus = []
for i in range(0, len(final_lst)):
    text = re.sub('[^a-zA-Z]', '', final_lst['Details'][i])
    text = text.lower()
    text = text.split()
    ps = PorterStemmer()
    text = ''.join(text)
    corpus.append(text)
cv = CountVectorizer(max_features=1500)
X = cv.fit_transform(corpus).toarray()
y = final_lst.iloc[:, 1].values

# spliting data for training and testing

print("spliting data for training and testing")
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)
classifier = GaussianNB();
classifier.fit(X_train, y_train)
acc_rate = accuracy_score(y_test, classifier.predict(X_test)) * 100
print("Groot has cross validated trained model and it is " + str(acc_rate) + "% accurate")

# save the trained model

print("saving the trained model")
f = open('trained_model.pickle', 'wb')
pickle.dump(classifier, f)
f.close()

################################################

# Read input Excel file

print("Reading input Excel file for prediction")
path = "D:\\Applens_Hackathon\\data for training.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
final_lstt = []
for i in range(2, 751):
    x = ""
    lst = []
    for j in range(1, 4):
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    lst.append(x)
    final_lstt.append(lst)
final_lstt = pd.DataFrame(final_lstt)
final_lstt.columns = ["Details"]
corpus = []

# preparing efficient data for prediction

for i in range(0, len(final_lstt)):
    text = re.sub('[^a-zA-Z]', '', final_lstt['Details'][i])
    text = text.lower()
    text = text.split()
    ps = PorterStemmer()
    text = ''.join(text)
    corpus.append(text)

XX = cv.transform(corpus).toarray()

# Load already trained model

print("Loading already trained model")
f = open('trained_model.pickle', 'rb')
classifier = pickle.load(f)

# Predict the appname

print("Predicting the appname")
da = classifier.predict(XX)
new_ar = []
for i in da:
    new_ar.append(str(i))

# Export predicted appname for each line of data

k = 0
for i in range(1, 2):
    cell_obj = sheet_obj.cell(row=1, column=5)
    cell_obj.value = "1CAppname_Predicted"
wb_obj.save("D:\\Applens_Hackathon\\ticket_dump.xlsx")
for i in range(2, 501):
    for j in range(5, 6):
        cell_obj = sheet_obj.cell(row=i, column=j)
        cell_obj.value = new_ar[k]
    k += 1
wb_obj.save("D:\\Applens_Hackathon\\ticket_dump.xlsx")
print("Groot has predicted application name and updated in input dump")
f.close()
