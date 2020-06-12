import openpyxl
obb=openpyxl.load_workbook("E://testt.xlsx")
o=obb.active
col_name=[]
tp=o.max_row
o.cell(row=1,column=27).value='row_count'
j=1
for i in range(2,tp+1):
    o.cell(row=i,column=27).value=j
    j+=1
obb.save("E://testt.xlsx")
for i in range(1,28):
    col_name.append(o.cell(row=1,column=i).value)

joi=" varchar(100), ".join(col_name)+" varchar(100)"

val_name=[]
val_names=[]
for j in range(2,1001):
    val_name=[]
    for i in range(1,28):
        cx=str(o.cell(row=j, column=i).value)
        val_name.append(cx)
    val_name=tuple(val_name)
    val_names.append(val_name)
vari=(', '.join(['%s']*27))

import mysql.connector

mydb=mysql.connector.connect(host='localhost',user='root',passwd='root',database='test')

mycursor=mydb.cursor()
xx="create table if not exists qwerty ({temp})"
mycursor.execute(xx.format(temp=joi))
mycursor.execute("ALTER TABLE qwerty modify Work_Log varchar(6000)")
mycursor.execute("ALTER TABLE qwerty modify Notes varchar(2000)")
mycursor.execute("ALTER TABLE qwerty modify Support_Diary varchar(4000)")
mycursor.execute("ALTER TABLE qwerty modify Resolution varchar(2000)")
up='''insert into qwerty values (%s)''' %(vari)
mycursor.executemany(up, val_names)
mydb.commit()
wb = openpyxl.Workbook()
sheet = wb.active
mycursor.execute("select * from qwerty")
m=2
for i in mycursor:
    n=1
    for j in i:
        c1 = sheet.cell(row=m, column=n)
        c1.value=j
        n+=1
    m+=1

mycursor.execute("show columns from qwerty")
m=1
for i in mycursor:
    for j in i:
        c1 = sheet.cell(row=1, column=m)
        c1.value = j
        m += 1
        break
#wb.save("E://updated.xlsx")

