import pandas as pd
import openpyxl
import re
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from sklearn.feature_extraction.text import CountVectorizer
'''final_lst=[[' AVMPMO_L2 AVM PMO Reports issue in AVM PMO', 'AVMPMO'],
           [' L1 (AVM Tools & Portals) AppLens Unable to view the Account name in the drop down', 'Applens'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet', 'AppLensLite'],
           [' L1 (AVM Tools & Portals) AppLensLite Queries on Debt configuration', 'AppLensLite'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet', 'AppLensLite'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet', 'AppLensLite'],
           [' AppLensLite_L2 AVM 1Capp Reports', 'AVM 1Capp'],
           [' L1 (AVM Tools & Portals) AppLens How to update Timesheet Approver', 'AppLensLite'],
           [' AVMPMO_L2 AVM PMO Reports issue in AVM PMO', 'AVMPMO'],
           [' L1 (AVM Tools & Portals) AppLens Unable to view the Account name in the drop down', 'Applens'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet', 'AppLensLite'],
           [' L1 (AVM Tools & Portals) AppLensLite Queries on Debt configuration', 'AppLensLite'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet', 'AppLensLite'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet', 'AppLensLite'],
           [' AppLensLite_L2 AVM 1Capp Reports', 'AVM 1Capp'],
           [' L1 (AVM Tools & Portals) AppLens How to update Timesheet Approver', 'AppLensLite']]'''
path = "E:\\testt_updated_v1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
final_lst = []
for i in range(2, 500):
    x = ""
    lst = []
    for j in range(1, 4):
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    lst.append(x)
    for j in range(9, 10):
        cell_obj = sheet_obj.cell(row=i, column=j)
        lst.append(cell_obj.value)
    final_lst.append(lst)

print(final_lst)


final_lst = pd.DataFrame(final_lst)
final_lst.columns = ["Details", "Appname"]
corpus = []
for i in range(0, len(final_lst)):
    text = re.sub('[^a-zA-Z]', '', final_lst['Details'][i])
    text = text.lower()
    text = text.split()
    ps = PorterStemmer()
    text = ''.join(text)
    corpus.append(text)
cv = CountVectorizer(max_features=1500)
X = cv.fit_transform(corpus).toarray()
y = final_lst.iloc[:, 1].values
from sklearn.model_selection import train_test_split
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.2, random_state = 0)

from sklearn.naive_bayes import GaussianNB

classifier = GaussianNB();
classifier.fit(X_train, y_train)


'''final_lstt=[[' AVMPMO_L2 AVM PMO Reports issue in AVM PMO'],
           [' L1 (AVM Tools & Portals) AppLens Unable to view the Account name in the drop down'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet'],
           [' L1 (AVM Tools & Portals) AppLensLite Queries on Debt configuration'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet'],
           [' L1 (AVM Tools & Portals) AppLensLite How to unfreeze timesheet'],
           [' dfgdgd fgfd'],
           [' L1 (AVM Tools & Portals) AppLens How to update Timesheet Approver']]'''


path = "E:\\testt_updated_v1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
final_lstt = []
for i in range(502, 1000):
    x = ""
    lst = []
    for j in range(1, 4):
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    lst.append(x)
    final_lstt.append(lst)

print(final_lstt)
final_lstt = pd.DataFrame(final_lstt)
final_lstt.columns = ["Details"]
    # nltk.download('stopwords')
corpus = []
for i in range(0, len(final_lstt)):
    text = re.sub('[^a-zA-Z]', '', final_lstt['Details'][i])
    text = text.lower()
    text = text.split()
    ps = PorterStemmer()
    text = ''.join(text)
    corpus.append(text)
print(corpus)

cv = CountVectorizer(max_features=1500)

X = cv.fit_transform(corpus).toarray()

print(classifier.predict(X))
