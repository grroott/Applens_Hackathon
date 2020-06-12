import openpyxl

obb = openpyxl.load_workbook("E://test_updated.xlsx")
o = obb.active
col_name = []
tp = o.max_row
o.cell(row=1, column=14).value = 'row_count'
j = 1
for i in range(2, tp + 1):
    o.cell(row=i, column=14).value = j
    j += 1
obb.save("E://test_updated.xlsx")
for i in range(1, 15):
    col_name.append(o.cell(row=1, column=i).value)

joi = " varchar(100), ".join(col_name) + " varchar(10)"

val_name = []
val_names = []
for j in range(2, 1001):
    val_name = []
    for i in range(1, 15):
        cx = str(o.cell(row=j, column=i).value)
        val_name.append(cx)
    val_name = tuple(val_name)
    val_names.append(val_name)

vari = (', '.join(['%s'] * 14))

import mysql.connector

mydb = mysql.connector.connect(host='localhost', user='root', passwd='root', database='test')

mycursor = mydb.cursor()
xx = "create table if not exists newtable ({temp})"
mycursor.execute(xx.format(temp=joi))
mycursor.execute("ALTER TABLE newtable modify Work_Log varchar(6000)")
mycursor.execute("ALTER TABLE newtable modify Notes varchar(2000)")
mycursor.execute("ALTER TABLE newtable modify Support_Diary varchar(4000)")
mycursor.execute("ALTER TABLE newtable modify Resolution varchar(2000)")
up = '''insert into newtable values (%s)''' % (vari)
mycursor.executemany(up, val_names)
mydb.commit()

wb = openpyxl.Workbook()
sheet = wb.active
mycursor.execute("select * from newtable")
m = 2
for i in mycursor:
    n = 1
    for j in i:
        c1 = sheet.cell(row=m, column=n)
        c1.value = j
        n += 1
    m += 1

mycursor.execute("show columns from newtable")
m = 1
for i in mycursor:
    for j in i:
        c1 = sheet.cell(row=1, column=m)
        c1.value = j
        m += 1
        break
wb.save("E://updated.xlsx")

