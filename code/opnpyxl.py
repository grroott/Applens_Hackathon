import openpyxl
import pandas as pd
import numpy as np
import re
import pickle
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from sklearn.feature_extraction.text import CountVectorizer

path="E:\\testt_updated_v1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

final_lst=[]
for i in range(2,10):
    x = ""
    lst = []
    for j in range(1,3):
        cell_obj = sheet_obj.cell(row = i, column = j)
        x=x+" "+str(cell_obj.value)
    lst.append(x)
    for j in range(9,10):
        cell_obj = sheet_obj.cell(row=i, column=j)
        lst.append(cell_obj.value)
    final_lst.append(lst)
print(final_lst)


final_lst = pd.DataFrame(final_lst)
final_lst.columns = ["Details", "Appname"]

print(len(final_lst))
# nltk.download('stopwords')
corpus = []
for i in range(0, len(final_lst)):
    text = re.sub('[^a-zA-Z]', '', final_lst['Details'][i])
    text = text.lower()
    text = text.split()
    ps = PorterStemmer()
    text = ''.join(text)
    # print(text)
    corpus.append(text)
print(corpus)
cv = CountVectorizer(max_features=1500)
X = cv.fit_transform(corpus).toarray()
y = final_lst.iloc[:, 1].values
# splitting the data set into training set and test set
from sklearn.model_selection import train_test_split
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.33)
# fitting naive bayes to the training set
from sklearn.naive_bayes import GaussianNB
classifier = GaussianNB();
classifier.fit(X_train, y_train)
f = open('my_classifier1.pickle', 'wb')
pickle.dump(classifier, f)

from sklearn.metrics import accuracy_score
print(accuracy_score(y_test,classifier.predict(X_test)))
print(classifier.predict(X_test))

print(y_test)

f.close()

