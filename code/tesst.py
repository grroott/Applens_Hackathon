import pandas as pd
import openpyxl
import re
import nltk
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from sklearn.feature_extraction.text import CountVectorizer
import pickle
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import accuracy_score
import pickle

print("Reading Excel for training the model...")

path = "E:\\data for training.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

max_row_in_train = sheet_obj.max_row
max_column_in_train = sheet_obj.max_column

main_data = []
bp_data = []
app_data = []
#    if cell_obj.value == 'Assigned to Group' or cell_obj.value == 'Category' or cell_obj.value == 'Item':
col_n = ['Assigned to Group', 'Category', 'Item']
for j in range(1, max_column_in_train+1):
    cell_obj = sheet_obj.cell(row=1, column=j)
    if cell_obj.value in col_n:
        main_data.append(j)
    elif cell_obj.value == 'Type':
        bp_data.append(j)
    elif cell_obj.value == '1CApp Name':
        app_data.append(j)

print(main_data)
bp_finaldata = []
app_finaldata = []
for i in range(2, max_row_in_train+1):
    x = ""
    common = []
    usecase1 = []
    usecase2 = []
    for j in main_data:
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    usecase1.append(x)
    usecase2.append(x)
    for j in bp_data:
        cell_obj = sheet_obj.cell(row=i, column=j)
        usecase1.append(cell_obj.value)
    for j in app_data:
        cell_obj = sheet_obj.cell(row=i, column=j)
        usecase2.append(cell_obj.value)
    bp_finaldata.append(usecase1)
    app_finaldata.append(usecase2)

wb_obj.close()

app_finaldata = pd.DataFrame(app_finaldata)
app_finaldata.columns = ["Details", "Appname"]
bp_finaldata = pd.DataFrame(bp_finaldata)
bp_finaldata.columns = ["Details", "Business_process"]

print("preparing efficient data for training...")


def preparing_data(inputt):
    prepared_data = []
    for ii in range(0, len(inputt)):
        m = re.sub('[^a-zA-Z]', '', inputt['Details'][ii])
        m = m.lower()
        m = m.split()
        # ps = PorterStemmer()
        m = ''.join(m)
        prepared_data.append(m)
    return prepared_data


app_latestdata = preparing_data(app_finaldata)
bp_latestdata = preparing_data(bp_finaldata)

print("spliting data for training and testing...")

cv = CountVectorizer(max_features=1500)
appname_data = cv.fit_transform(app_latestdata).toarray()
appname = app_finaldata.iloc[:, 1].values
bpname_data = cv.fit_transform(bp_latestdata).toarray()
bpname = bp_finaldata.iloc[:, 1].values

appdata1_train, appdata1_test, appdata2_train, appdata2_test = train_test_split(appname_data, appname, test_size=0.33)
classifier1 = GaussianNB()
classifier1.fit(appdata1_train, appdata2_train)
acc_rate = accuracy_score(appdata2_test, classifier1.predict(appdata1_test)) * 100

# print(classifier1.predict(appdata1_test))
print("######################")
print("Groot has cross validated trained model for Appname and it is " + str(round(acc_rate, 2)) + "% accurate")

bpdata1_train, bpdata1_test, bpdata2_train, bpdata2_test = train_test_split(bpname_data, bpname, test_size=0.33)
classifier2 = GaussianNB()
classifier2.fit(bpdata1_train, bpdata2_train)
acc_rate = accuracy_score(bpdata2_test, classifier2.predict(bpdata1_test)) * 100

# print(classifier2.predict(bpdata1_test))
print("Groot has cross validated trained model for Business process and it is "
      + str(round(acc_rate, 2)) + "% accurate")
print("######################")

print("saving the trained model")
f1 = open('trained_model_app.pickle', 'wb')
f2 = open('trained_model_bp.pickle', 'wb')
pickle.dump(classifier1, f1)
pickle.dump(classifier2, f2)
f1.close()
f2.close()

#################################################################
print("Reading input Excel file for prediction...")
path = "E:\\data for pred.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

max_row_in_test = sheet_obj.max_row
max_column_in_test = sheet_obj.max_row

prep_data = []
for j in range(1, max_column_in_test+1):
    cell_obj = sheet_obj.cell(row=1, column=j)
    if cell_obj.value == 'Assigned to Group' or cell_obj.value == 'Category' or cell_obj.value == 'Item':
        prep_data.append(j)

final_lstt = []
for i in range(2, max_row_in_test+1):
    x = ""
    lst = []
    for j in prep_data:
        cell_obj = sheet_obj.cell(row=i, column=j)
        x = x + " " + str(cell_obj.value)
    lst.append(x)
    final_lstt.append(lst)

final_lstt = pd.DataFrame(final_lstt)
final_lstt.columns = ["Details"]

corpus = preparing_data(final_lstt)

XX = cv.transform(corpus).toarray()

print("Loading already trained model")
f1 = open('trained_model_app.pickle', 'rb')
f2 = open('trained_model_bp.pickle', 'rb')
classifie1 = pickle.load(f1)
classifie2 = pickle.load(f2)

print("Predicting the Appname...")
da = classifie1.predict(XX)
new_ar = []
for i in da:
    new_ar.append(str(i))
print("Predicting the Business process...")
daa = classifie2.predict(XX)
new_arr = []
for i in daa:
    new_arr.append(str(i))

# Export predicted appname for each line of data
print("Exporting predicted results to the dump...")
sheet_obj.cell(row=1, column=7).value = "1CAppname_Predicted"
sheet_obj.cell(row=1, column=8).value = "BP_Predicted"

tmp1 = 0
for i in range(2, 1002):
    for j in range(7, 8):
        cell_obj = sheet_obj.cell(row=i, column=j).value = new_ar[tmp1]
    tmp1 += 1
tmp2 = 0
for i in range(2, 1002):
    for j in range(8, 9):
        cell_obj = sheet_obj.cell(row=i, column=j).value = new_arr[tmp2]
    tmp2 += 1

wb_obj.save("E:\\data for pred.xlsx")
print("Groot has predicted Application name, Business process and updated the input dump!!!")
wb_obj.close()
